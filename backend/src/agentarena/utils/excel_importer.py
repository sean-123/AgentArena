"""Excel import for testcases."""

import json
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


def import_excel_to_testcases(content: bytes) -> list[dict[str, Any]]:
    """
    Import testcases from Excel file.
    Expected columns: id, question, persona_question, key_points, domain, difficulty
    key_points can be JSON array string or comma-separated.
    """
    wb = load_workbook(BytesIO(content), read_only=True)
    ws = wb.active
    if not ws:
        return []
    headers = []
    testcases = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx == 0:
            headers = [str(h).strip().lower() if h else "" for h in row]
            continue
        row_dict = dict(zip(headers, row))
        if not any(v for v in row_dict.values() if v):
            continue
        # Map common column names
        question = (
            row_dict.get("question")
            or row_dict.get("q")
            or row_dict.get("问题")
            or ""
        )
        if not question:
            continue
        key_points_raw = (
            row_dict.get("key_points")
            or row_dict.get("keypoints")
            or row_dict.get("要点")
            or row_dict.get("关键点")
        )
        # 当 key_points 列为空时，若 persona_question 为逗号分隔的要点格式，则作为 key_points 解析
        persona_used_as_key_points = False
        if not key_points_raw:
            pq = row_dict.get("persona_question") or row_dict.get("persona")
            if pq and isinstance(pq, str) and "," in pq.strip():
                key_points_raw = pq
                persona_used_as_key_points = True
        key_points = []
        if key_points_raw:
            if isinstance(key_points_raw, str):
                if key_points_raw.strip().startswith("["):
                    try:
                        key_points = json.loads(key_points_raw)
                    except json.JSONDecodeError:
                        key_points = [x.strip() for x in key_points_raw.split(",") if x.strip()]
                else:
                    key_points = [x.strip() for x in key_points_raw.split(",") if x.strip()]
            elif isinstance(key_points_raw, (list, tuple)):
                key_points = list(key_points_raw)
        persona_question_val = None
        if not persona_used_as_key_points:
            persona_question_val = _cell_str(row_dict.get("persona_question") or row_dict.get("persona"))
        testcases.append({
            "id": _cell_str(row_dict.get("id")),
            "question": str(question),
            "persona_question": persona_question_val,
            "key_points": key_points if key_points else None,
            "domain": _cell_str(row_dict.get("domain")),
            "difficulty": _cell_str(row_dict.get("difficulty")),
        })
    return testcases


def _cell_str(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None
