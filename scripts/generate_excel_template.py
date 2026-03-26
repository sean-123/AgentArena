#!/usr/bin/env python3
"""生成测试用例导入的 Excel 模板。"""

import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def main():
    wb = Workbook()
    ws = wb.active
    if not ws:
        return
    ws.title = "测试用例"

    # 表头（支持中英文）
    headers = [
        ("id", "ID（可选）"),
        ("question", "问题（必填）"),
        ("persona_question", "角色问题（可选）"),
        ("key_points", "要点（可选，逗号分隔或 JSON 数组）"),
        ("domain", "领域（可选）"),
        ("difficulty", "难度（可选）"),
    ]
    for col, (en, _) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=en)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # 示例数据
    examples = [
        [
            "tc_001",
            "如何实现一个简单的 REST API？",
            "假设你是后端工程师，请说明",
            "RESTful,HTTP,API设计",
            "后端开发",
            "简单",
        ],
        [
            "tc_002",
            "解释什么是依赖注入",
            "",
            '["依赖注入", "设计模式", "解耦"]',
            "软件设计",
            "中等",
        ],
    ]
    for row_idx, row_data in enumerate(examples, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 自动列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # 保存到 templates 和 frontend/public
    root = Path(__file__).resolve().parent.parent
    (root / "templates").mkdir(parents=True, exist_ok=True)
    out_path = root / "templates" / "testcase-import-template.xlsx"
    wb.save(out_path)
    print(f"已生成: {out_path}")
    # 同步到前端供下载
    frontend_public = root / "frontend" / "public"
    frontend_public.mkdir(parents=True, exist_ok=True)
    shutil.copy(out_path, frontend_public / "testcase-import-template.xlsx")
    print(f"已同步: {frontend_public / 'testcase-import-template.xlsx'}")


if __name__ == "__main__":
    main()
